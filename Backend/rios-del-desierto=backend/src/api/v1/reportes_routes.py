from flask import Blueprint, send_file, jsonify
from io import BytesIO
from datetime import datetime, timedelta
from sqlalchemy import select, func

import pandas as pd

from src.models.cliente import Cliente
from src.models.compra import Compra
from src.models.enums import EstadoCompraEnum
from src.extensions import db

bp = Blueprint("reportes", __name__)


@bp.get("/reportes/clientes-fidelizacion")
def generar_reporte_clientes_fidelizacion():
    """
    Genera un reporte en Excel con los productos comprados por clientes que superan 5'000.000 COP
    en compras del último mes. Cada producto comprado se muestra en una fila separada.
    
    El reporte incluye:
        - Datos básicos del cliente (nombre, apellido, correo, teléfono)
        - Tipo y número de documento
        - Fecha de la compra
        - Nombre del producto
        - Cantidad comprada
        - Precio unitario
        - Subtotal (cantidad * precio unitario)
    
    Returns:
        Archivo Excel (.xlsx) con el reporte
    """
    try:
        # Obtener todos los detalles de compra con productos usando el join
        detalles_compras = Compra.obtener_detalles_compras_con_productos_ultimo_mes(
            monto_minimo_total=5_000_000
        )
        
        if not detalles_compras:
            return jsonify({
                "message": "No hay clientes que cumplan el criterio de fidelización (monto > 5'000.000 COP en el último mes)"
            }), 404
        
        # Preparar datos para el DataFrame
        datos_reporte = []
        
        for compra, detalle, producto in detalles_compras:
            # Obtener información del cliente
            cliente = compra.cliente
            
            # Obtener información del documento
            tipo_documento = cliente.documento.tipo_documento.value if cliente.documento else "N/A"
            numero_documento = cliente.documento.numero_documento if cliente.documento else "N/A"
            
            # Formatear fecha de compra
            fecha_str = compra.fecha.strftime("%Y-%m-%d %H:%M:%S")
            
            # Calcular subtotal
            subtotal = detalle.cantidad_compra * detalle.precio_unitario
            
            datos_reporte.append({
                "Nombre": cliente.nombre,
                "Apellido": cliente.apellido,
                "Correo Electrónico": cliente.correo_electronico,
                "Teléfono": cliente.telefono_celular,
                "Tipo Documento": tipo_documento,
                "Número Documento": numero_documento,
                "Fecha Compra": fecha_str,
                "Producto": producto.nombre,
                "Cantidad": detalle.cantidad_compra,
                "Precio Unitario (COP)": detalle.precio_unitario,
                "Subtotal (COP)": subtotal,
            })
        
        # Crear DataFrame
        df = pd.DataFrame(datos_reporte)
        
        if df.empty:
            return jsonify({
                "message": "No hay productos que cumplan el criterio de fidelización"
            }), 404
        
        # Ordenar por cliente y luego por fecha de compra descendente
        df = df.sort_values(["Nombre", "Apellido", "Fecha Compra"], ascending=[True, True, False])
        
        # Agrupar por cliente y agregar filas de total por cliente
        filas_finales = []
        filas_totales_indices = []  # Para formatear después (índices en el DataFrame final)
        fila_actual = 0
        
        # Agrupar por cliente (usando Nombre + Apellido como identificador único)
        for (nombre, apellido), grupo in df.groupby(["Nombre", "Apellido"]):
            # Agregar todas las filas del cliente
            filas_finales.append(grupo)
            fila_actual += len(grupo)
            
            # Calcular el total del cliente
            total_cliente = grupo["Subtotal (COP)"].sum()
            
            # Crear fila de total para este cliente
            fila_total = pd.DataFrame([{
                "Nombre": "",
                "Apellido": "",
                "Correo Electrónico": "",
                "Teléfono": "",
                "Tipo Documento": "",
                "Número Documento": "",
                "Fecha Compra": "",
                "Producto": f"TOTAL {nombre} {apellido}",
                "Cantidad": "",
                "Precio Unitario (COP)": "",
                "Subtotal (COP)": total_cliente,
            }])
            
            filas_finales.append(fila_total)
            filas_totales_indices.append(fila_actual)  # Índice de la fila de total en el DataFrame
            fila_actual += 1
        
        # Concatenar todas las filas
        df_final = pd.concat(filas_finales, ignore_index=True)
        
        # Crear archivo Excel en memoria
        output = BytesIO()

        # Escribir el DataFrame en el archivo Excel:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Escribir el DataFrame empezando en la fila 2 (fila 1 es el header, fila 2 son los encabezados)
            df_final.to_excel(writer, sheet_name='Clientes Fidelización', index=False, startrow=1)
            
            # Obtener la hoja para formatear
            worksheet = writer.sheets['Clientes Fidelización']
            
            # Agregar header con título grande ANTES de los encabezados
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Insertar una fila al inicio para el header
            worksheet.insert_rows(1)
            
            # Obtener el número de columnas
            num_columnas = len(df_final.columns)
            
            # Combinar celdas para el header (de la columna 1 a la última)
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columnas)
            
            # Agregar el texto del header
            header_cell = worksheet.cell(row=1, column=1)
            header_cell.value = "Reporte de Fidelización Clientes - Rios del Desierto S.A.S."
            header_cell.font = Font(size=18, bold=True)
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Ajustar altura de la fila del header
            worksheet.row_dimensions[1].height = 30
            
            # Ajustar ancho de columnas
            for idx, col in enumerate(df_final.columns, 1):
                max_length = max(
                    df_final[col].astype(str).map(len).max(),
                    len(col)
                ) + 2
                col_letter = get_column_letter(idx)
                worksheet.column_dimensions[col_letter].width = min(max_length, 50)
            
            # Formatear las filas de total por cliente
            # +2 porque: 1 fila header, 1 fila encabezados de columnas
            fila_inicio_datos = 4  # Fila 1: header, Fila 2: encabezados, Fila 3: datos
            
            # Aplicar formato a todas las filas de total por cliente
            for idx_total in filas_totales_indices:
                fila_excel = fila_inicio_datos + idx_total
                for col_idx in range(1, len(df_final.columns) + 1):
                    cell = worksheet.cell(row=fila_excel, column=col_idx)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        output.seek(0)
        
        # Generar nombre del archivo con fecha
        fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"reporte_clientes_fidelizacion_{fecha_str}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nombre_archivo
        )
        
    except Exception as e:
        return jsonify({
            "error": "Error al generar el reporte",
            "message": str(e)
        }), 500
